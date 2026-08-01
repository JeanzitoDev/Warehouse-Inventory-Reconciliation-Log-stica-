import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
import config
from database import fetch_system_inventory

def reconcile_inventory(filename: str):
    """Lê a planilha física, compara com o DB e gera arquivo formatado de divergências."""
    print(f"Iniciando reconciliação para: {filename}")
    
    input_path = config.INPUT_DIR / filename
    if not input_path.exists():
        print(f"❌ Arquivo não encontrado: {input_path}")
        return

    # 1. Carrega Contagem Física (Raw Data)
    df_physical = pd.read_excel(input_path)
    
    # 2. Carrega Estoque do Sistema (DB)
    df_system = fetch_system_inventory()
    
    if df_system.empty:
        print("⚠️ DB retornou vazio. Usando dados mockados para demonstração.")
        # Mock de dados caso rode sem banco
        df_system = pd.DataFrame([
            {'sku': 'SKU-001', 'system_qty': 100},
            {'sku': 'SKU-002', 'system_qty': 50}
        ])

    # 3. Merge & Lógica de Negócio (Reconciliação)
    df_merged = pd.merge(df_physical, df_system, on='sku', how='left')
    df_merged['system_qty'] = df_merged['system_qty'].fillna(0)
    df_merged['discrepancy'] = df_merged['physical_qty'] - df_merged['system_qty']
    
    # Define Status
    def define_status(row):
        if row['discrepancy'] == 0: return 'OK'
        if row['discrepancy'] < 0: return 'MISSING'
        return 'SURPLUS'
        
    df_merged['status'] = df_merged.apply(define_status, axis=1)
    
    # 4. Geração de Relatório Formatado (Openpyxl)
    output_path = config.OUTPUT_DIR / f"Reconciled_{filename}"
    df_merged.to_excel(output_path, index=False, sheet_name="Reconciliation")
    
    format_excel_report(output_path)
    print(f"✅ Reconciliação salva em: {output_path.name}")

def format_excel_report(filepath: Path):
    """Aplica formatação condicional baseada nas regras de negócio (Openpyxl)."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Reconciliation"]
    
    # Formata Cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Formatação Condicional das Linhas
    status_col_idx = 5 # Supondo que 'status' é a 5ª coluna
    
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=status_col_idx).value
        
        if status in config.DISCREPANCY_RULES:
            color_hex = config.DISCREPANCY_RULES[status]['color']
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

    wb.save(filepath)

if __name__ == "__main__":
    reconcile_inventory("physical_count_week12.xlsx")